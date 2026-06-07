"""
Generador de la corrección de calendario de fase de grupos (Mundial 2026).

Fuente de verdad: el Excel oficial (local, visitante, fecha, hora CDMX, estadio, ciudad).
Estado actual: se reproduce de forma determinista el calendario sintetico del seed
(round-robin + formula de kickoff), que es lo que hay hoy en prod.

Emite 3 artefactos en quiniela-2026-familia/scripts/:
  1) reporte-comparativo-partidos.xlsx  -> Fase A1 (solo lectura, para revisar)
  2) fix-group-schedule.sql             -> Fase A2 + A3 (transaccional, idempotente)
  3) seed-group-matches.sql             -> bloque VALUES para reescribir seed.sql

NO toca la base de datos. Solo lee el Excel y escribe archivos.
"""

from __future__ import annotations

import os
from datetime import datetime, timedelta

import openpyxl

# ---------------------------------------------------------------------------
# Rutas
# ---------------------------------------------------------------------------
EXCEL_PATH = r"C:\Users\alber\Downloads\Mundial_2026_Fase_Grupos_1.xlsx"
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_XLSX = os.path.join(SCRIPTS_DIR, "reporte-comparativo-partidos.xlsx")
FIX_SQL = os.path.join(SCRIPTS_DIR, "fix-group-schedule.sql")
SEED_SQL = os.path.join(SCRIPTS_DIR, "seed-group-matches.sql")

# ---------------------------------------------------------------------------
# Mapa de equipos: (grupo, nombre canonico BD) -> code  (migraciones 003/004)
# ---------------------------------------------------------------------------
TEAMS = {
    "A": ["México", "Sudáfrica", "Corea del Sur", "República Checa"],
    "B": ["Canadá", "Bosnia y Herzegovina", "Qatar", "Suiza"],
    "C": ["Brasil", "Marruecos", "Escocia", "Haití"],
    "D": ["Estados Unidos", "Paraguay", "Turquía", "Australia"],
    "E": ["Alemania", "Costa de Marfil", "Ecuador", "Curazao"],
    "F": ["Países Bajos", "Suecia", "Túnez", "Japón"],
    "G": ["Bélgica", "Egipto", "Irán", "Nueva Zelanda"],
    "H": ["España", "Arabia Saudita", "Uruguay", "Cabo Verde"],
    "I": ["Francia", "Senegal", "Irak", "Noruega"],
    "J": ["Argentina", "Argelia", "Austria", "Jordania"],
    "K": ["Portugal", "RD del Congo", "Uzbekistán", "Colombia"],
    "L": ["Inglaterra", "Croacia", "Ghana", "Panamá"],
}

# code -> nombre, y nombre -> code
CODE_TO_NAME: dict[str, str] = {}
NAME_TO_CODE: dict[str, str] = {}
for g, names in TEAMS.items():
    for i, name in enumerate(names, start=1):
        code = f"E{g}{i}"
        CODE_TO_NAME[code] = name
        NAME_TO_CODE[name] = code

# Alias: nombre como viene en el Excel -> nombre canonico BD
ALIASES = {
    "EE.UU.": "Estados Unidos",
    "Rep. Checa": "República Checa",
    "Bosnia-Herzegovina": "Bosnia y Herzegovina",
    "Rep. Dem. del Congo": "RD del Congo",
}

GROUP_IDX = {g: i for i, g in enumerate("ABCDEFGHIJKL")}  # A=0 .. L=11

# par de indices (ordenado) -> numero de partido round-robin
PAIR_TO_MATCHNUM = {(1, 2): 1, (3, 4): 2, (1, 3): 3, (2, 4): 4, (1, 4): 5, (2, 3): 6}
# numero de partido -> (idx local, idx visitante) que usa HOY el seed
MATCHNUM_TO_PATTERN = {1: (1, 2), 2: (3, 4), 3: (1, 3), 4: (2, 4), 5: (1, 4), 6: (2, 3)}

MONTHS = {"Ene": 1, "Feb": 2, "Mar": 3, "Abr": 4, "May": 5, "Jun": 6,
          "Jul": 7, "Ago": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dic": 12}


def code_for(group: str, excel_name: str) -> str:
    name = ALIASES.get(excel_name.strip(), excel_name.strip())
    code = NAME_TO_CODE.get(name)
    if not code:
        raise ValueError(f"No encuentro code para '{excel_name}' (grupo {group})")
    if code[1] != group:
        raise ValueError(f"'{name}' ({code}) no pertenece al grupo {group} del Excel")
    return code


def parse_date(fecha: str, hora: str) -> tuple[int, int, int]:
    """'Jueves, 11 Jun' + '13:00' -> (dia, hh, mm). Mes siempre Jun (2026)."""
    part = fecha.split(",")[-1].strip()           # '11 Jun'
    day_str, mon_str = part.split()
    day = int(day_str)
    mon = MONTHS[mon_str[:3].capitalize()]
    assert mon == 6, f"Mes inesperado en '{fecha}'"
    hh, mm = (int(x) for x in str(hora).strip().split(":"))
    return day, hh, mm


def current_kickoff_utc(g_idx: int, match_num: int) -> datetime:
    """Reproduce la formula del seed sintetico (en UTC)."""
    base = datetime(2026, 6, 11, 12, 0, 0)
    do = 0 if match_num in (1, 2) else (6 if match_num in (3, 4) else 12)
    do += g_idx // 2
    slot = (0 if match_num % 2 == 1 else 1) + (g_idx % 2) * 2
    return base + timedelta(days=do, hours=slot * 3)


def sql_str(s: str) -> str:
    return s.replace("'", "''")


def fmt_cdmx(dt_utc: datetime) -> str:
    return (dt_utc - timedelta(hours=6)).strftime("%d-%b %H:%M")


# ---------------------------------------------------------------------------
# Leer Excel -> filas oficiales
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb["Fase de Grupos"]

records = []
for row in ws.iter_rows(values_only=True):
    if not row or not isinstance(row[0], int):
        continue
    num, grupo, local, visit, fecha, hora, estadio, ciudad = row[:8]
    grupo = str(grupo).strip()
    home_code = code_for(grupo, local)
    away_code = code_for(grupo, visit)
    h_idx, a_idx = int(home_code[2]), int(away_code[2])
    pair = tuple(sorted((h_idx, a_idx)))
    match_num = PAIR_TO_MATCHNUM[pair]
    match_id = f"g_{grupo}_{match_num}"

    day, hh, mm = parse_date(fecha, hora)
    kickoff_sql = f"2026-06-{day:02d} {hh:02d}:{mm:02d}:00-06"  # CDMX (UTC-6)
    kickoff_utc = datetime(2026, 6, day, hh, mm) + timedelta(hours=6)

    # Estado actual (seed sintetico)
    g_idx = GROUP_IDX[grupo]
    cur_home_idx, cur_away_idx = MATCHNUM_TO_PATTERN[match_num]
    cur_home_code = f"E{grupo}{cur_home_idx}"
    cur_away_code = f"E{grupo}{cur_away_idx}"
    cur_kick_utc = current_kickoff_utc(g_idx, match_num)

    order_wrong = home_code != cur_home_code
    date_wrong = kickoff_utc != cur_kick_utc

    records.append({
        "num": num, "grupo": grupo, "match_id": match_id,
        "home_code": home_code, "away_code": away_code,
        "cur_home_code": cur_home_code, "cur_away_code": cur_away_code,
        "kickoff_sql": kickoff_sql, "kickoff_utc": kickoff_utc,
        "cur_kick_utc": cur_kick_utc,
        "estadio": str(estadio).strip(), "ciudad": str(ciudad).strip(),
        "order_wrong": order_wrong, "date_wrong": date_wrong,
    })

records.sort(key=lambda r: (r["grupo"], int(r["match_id"].split("_")[2])))
assert len(records) == 72, f"Esperaba 72 partidos, encontre {len(records)}"
ids = {r["match_id"] for r in records}
assert len(ids) == 72, "Hay match_id duplicados — revisar mapeo"

n_date = sum(r["date_wrong"] for r in records)
n_order = sum(r["order_wrong"] for r in records)

# ---------------------------------------------------------------------------
# 1) Reporte comparativo (xlsx)
# ---------------------------------------------------------------------------
out = openpyxl.Workbook()
sh = out.active
sh.title = "Comparativo"
headers = ["# oficial", "match_id", "Grupo",
           "Local ACTUAL", "Visit. ACTUAL", "Fecha/hora ACTUAL (CDMX)",
           "Local OFICIAL", "Visit. OFICIAL", "Fecha/hora OFICIAL (CDMX)",
           "Estadio", "Ciudad",
           "¿Fecha/hora mal?", "¿Orden mal?", "Acción"]
sh.append(headers)
for r in records:
    if r["order_wrong"] and r["date_wrong"]:
        accion = "Fecha + Orden"
    elif r["order_wrong"]:
        accion = "Solo orden"
    elif r["date_wrong"]:
        accion = "Solo fecha/hora"
    else:
        accion = "OK"
    sh.append([
        r["num"], r["match_id"], r["grupo"],
        CODE_TO_NAME[r["cur_home_code"]], CODE_TO_NAME[r["cur_away_code"]],
        fmt_cdmx(r["cur_kick_utc"]),
        CODE_TO_NAME[r["home_code"]], CODE_TO_NAME[r["away_code"]],
        fmt_cdmx(r["kickoff_utc"]),
        r["estadio"], r["ciudad"],
        "SÍ" if r["date_wrong"] else "no",
        "SÍ" if r["order_wrong"] else "no",
        accion,
    ])
widths = [8, 10, 7, 16, 16, 22, 16, 16, 22, 24, 18, 14, 12, 16]
for i, w in enumerate(widths, start=1):
    sh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
sh.freeze_panes = "A2"
out.save(REPORT_XLSX)

# ---------------------------------------------------------------------------
# 2) fix-group-schedule.sql  (A2 + A3, transaccional, idempotente)
# ---------------------------------------------------------------------------
lines = []
lines.append("-- Correccion one-time del calendario de fase de grupos (Mundial 2026).")
lines.append("-- Generado por scripts/generate-schedule-fix.py desde el Excel oficial.")
lines.append("-- Idempotente: kickoff/venue absolutos; swaps protegidos por orientacion actual.")
lines.append("--")
lines.append("-- USO (dos pasadas):")
lines.append("--   1) DRY-RUN: psql \"$env:SUPABASE_DB_URL\" -v ON_ERROR_STOP=1 -f scripts\\fix-group-schedule.sql")
lines.append("--      (termina en 'rollback;' -> no cambia nada, solo imprime la verificacion)")
lines.append("--   2) Si V1..V4 estan OK: cambiar la ultima linea 'rollback;' por 'commit;' y re-correr.")
lines.append("")
lines.append("begin;")
lines.append("")
lines.append("-- Snapshots para verificacion (se borran al cerrar la sesion).")
lines.append("create temp table _pred_before on commit drop as select * from predictions;")
lines.append("create temp table _match_before on commit drop as "
             "select id, home_team_code, away_team_code from matches;")
lines.append("")
lines.append("-- 0) Columnas de sede (aditivo, sin riesgo).")
lines.append("alter table matches add column if not exists stadium text;")
lines.append("alter table matches add column if not exists city    text;")
lines.append("")
lines.append("-- ===========================================================================")
lines.append("-- FASE A2: fecha, hora, estadio, ciudad (UPDATE absoluto, no toca pronosticos)")
lines.append("-- ===========================================================================")
for r in records:
    lines.append(
        f"update matches set kickoff_at='{r['kickoff_sql']}', "
        f"stadium='{sql_str(r['estadio'])}', city='{sql_str(r['ciudad'])}' "
        f"where id='{r['match_id']}';"
    )
lines.append("")
lines.append("-- ===========================================================================")
lines.append("-- FASE A3: localia volteada -> swap equipos + swap marcadores de pronosticos")
lines.append("-- (pronosticos PRIMERO, luego el partido; ambos con guard de orientacion).")
lines.append("-- ===========================================================================")
order_fixes = [r for r in records if r["order_wrong"]]
for r in order_fixes:
    cur_home = r["cur_home_code"]  # local actual (erroneo) = guard
    lines.append(f"-- {r['match_id']}: {CODE_TO_NAME[cur_home]} (local actual) -> "
                 f"{CODE_TO_NAME[r['home_code']]} (local oficial)")
    lines.append(
        f"update predictions p set home_score=p.away_score, away_score=p.home_score "
        f"from matches m where p.match_id=m.id and m.id='{r['match_id']}' "
        f"and m.home_team_code='{cur_home}';"
    )
    lines.append(
        f"update matches set home_team_code=away_team_code, away_team_code=home_team_code, "
        f"home_score=away_score, away_score=home_score "
        f"where id='{r['match_id']}' and home_team_code='{cur_home}';"
    )
lines.append("")
lines.append("-- ===========================================================================")
lines.append("-- VERIFICACION (se imprime ANTES del rollback/commit).")
lines.append("-- ===========================================================================")
lines.append("")
lines.append("-- (V1) Conteo de pronosticos intacto. Esperado: antes = despues.")
lines.append("select (select count(*) from _pred_before) as antes,")
lines.append("       (select count(*) from predictions)  as despues;")
lines.append("")
lines.append("-- (V2) Significado por identidad preservado. Esperado: filas = 0.")
lines.append("--      Cada (jugador, partido, equipo, goles) de antes existe despues y viceversa.")
lines.append("with before_unpivot as (")
lines.append("  select b.player_id, b.match_id, mb.home_team_code as team, b.home_score as goals")
lines.append("  from _pred_before b join _match_before mb on mb.id=b.match_id")
lines.append("  union all")
lines.append("  select b.player_id, b.match_id, mb.away_team_code, b.away_score")
lines.append("  from _pred_before b join _match_before mb on mb.id=b.match_id),")
lines.append("after_unpivot as (")
lines.append("  select p.player_id, p.match_id, m.home_team_code as team, p.home_score as goals")
lines.append("  from predictions p join matches m on m.id=p.match_id")
lines.append("  union all")
lines.append("  select p.player_id, p.match_id, m.away_team_code, p.away_score")
lines.append("  from predictions p join matches m on m.id=p.match_id)")
lines.append("select 'meaning_changed' as problema, count(*) as filas from (")
lines.append("  (select * from before_unpivot except select * from after_unpivot)")
lines.append("  union all")
lines.append("  (select * from after_unpivot except select * from before_unpivot)) d;")
lines.append("")
lines.append("-- (V3) Estado final == oficial (localia, kickoff, sede). Esperado: 0 filas.")
lines.append("with official(match_id, home_code, away_code, kickoff_at, stadium, city) as (values")
official_rows = []
for r in records:
    official_rows.append(
        f"  ('{r['match_id']}','{r['home_code']}','{r['away_code']}',"
        f"timestamptz '{r['kickoff_sql']}','{sql_str(r['estadio'])}','{sql_str(r['ciudad'])}')"
    )
lines.append(",\n".join(official_rows))
lines.append(")")
lines.append("select m.id, m.home_team_code, o.home_code, m.away_team_code, o.away_code, m.kickoff_at")
lines.append("from matches m join official o on o.match_id=m.id")
lines.append("where m.home_team_code<>o.home_code or m.away_team_code<>o.away_code")
lines.append("   or m.kickoff_at<>o.kickoff_at")
lines.append("   or m.stadium is distinct from o.stadium or m.city is distinct from o.city;")
lines.append("")
lines.append("-- (V4) Total de partidos de grupo. Esperado: 72.")
lines.append("select count(*) as total_group from matches where stage='group';")
lines.append("")
lines.append("-- ---------------------------------------------------------------------------")
lines.append("-- Si V1 (antes=despues), V2 (0), V3 (0) y V4 (72) -> cambiar la ultima linea a")
lines.append("-- 'commit;' y re-correr. Mientras diga 'rollback;' es dry-run (no cambia nada).")
lines.append("-- ---------------------------------------------------------------------------")
lines.append("rollback;  -- <<< DRY-RUN. Cambiar a 'commit;' para aplicar de verdad.")
lines.append("")

with open(FIX_SQL, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

# ---------------------------------------------------------------------------
# 3) seed-group-matches.sql  (bloque VALUES para reescribir seed.sql)
# ---------------------------------------------------------------------------
seed = []
seed.append("-- Partidos de fase de grupos (72) con datos OFICIALES del Mundial 2026.")
seed.append("-- Generado por scripts/generate-schedule-fix.py. Reemplaza el generador sintetico.")
seed.append("insert into matches "
            "(id, stage, group_letter, home_team_code, away_team_code, kickoff_at, stadium, city)")
seed.append("values")
vals = []
for r in records:
    vals.append(
        f"  ('{r['match_id']}','group','{r['grupo']}',"
        f"'{r['home_code']}','{r['away_code']}',"
        f"'{r['kickoff_sql']}','{sql_str(r['estadio'])}','{sql_str(r['ciudad'])}')"
    )
seed.append(",\n".join(vals))
seed.append("on conflict (id) do nothing;")
seed.append("")
with open(SEED_SQL, "w", encoding="utf-8") as f:
    f.write("\n".join(seed))

# ---------------------------------------------------------------------------
# Resumen
# ---------------------------------------------------------------------------
print(f"Partidos procesados : {len(records)}")
print(f"Fecha/hora mal      : {n_date}")
print(f"Localia volteada    : {n_order}")
print(f"Ambos               : {sum(r['order_wrong'] and r['date_wrong'] for r in records)}")
print(f"OK (sin cambios)    : {sum(not r['order_wrong'] and not r['date_wrong'] for r in records)}")
print()
print("Artefactos escritos:")
print(f"  {REPORT_XLSX}")
print(f"  {FIX_SQL}")
print(f"  {SEED_SQL}")
print()
print("Muestra (primeros 6 partidos del grupo A..):")
for r in records[:6]:
    print(f"  {r['match_id']}: actual {CODE_TO_NAME[r['cur_home_code']]} vs "
          f"{CODE_TO_NAME[r['cur_away_code']]} ({fmt_cdmx(r['cur_kick_utc'])})  ->  "
          f"oficial {CODE_TO_NAME[r['home_code']]} vs {CODE_TO_NAME[r['away_code']]} "
          f"({fmt_cdmx(r['kickoff_utc'])})  [{'F' if r['date_wrong'] else '-'}"
          f"{'O' if r['order_wrong'] else '-'}]")
